import requests
import pandas as pd
from datetime import datetime, date
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from openpyxl import Workbook

# Set up logging
logging.basicConfig(
    filename='price_puller.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def get_avg_price(url, sheet_name, output_file='Price_Puller.xlsx'):
    headers = {
        "accept": "*/*",
        "accept-language": "en-US,en;q=0.9",
        "content-type": "application/json",
        "priority": "u=1, i",
        "x-fwd-svc": "atc"
    }

    session = requests.Session()
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    session.mount('https://', HTTPAdapter(max_retries=retries))

    try:
        response = session.get(url, headers=headers)

        if response.status_code != 200:
            return

        data = response.json()
        links = data.get('links', [])

        today_date = date.today()
        today_str = today_date.strftime('%d%b%Y').upper()

        row = {'Date': today_str}
        price_columns = []

        for link in links:
            year = link.get('value')
            avg_price = link.get('avgPrice')
            if year and avg_price:
                year = int(year)
                row[year] = avg_price
                price_columns.append(year)

        # Calculate row number from anchor date (22Apr2025)
        anchor_date = datetime.strptime("22Apr2025", "%d%b%Y").date()
        delta_days = (today_date - anchor_date).days + 2  # Row 2 = anchor

        # Load or create workbook
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
        else:
            wb = Workbook()

        # Get or create the sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # Set headers in first row if missing
        existing_headers = [cell.value for cell in ws[1] if cell.value is not None]
        all_headers = ['Date'] + sorted(price_columns)

        if not existing_headers:
            for col_index, header in enumerate(all_headers, start=1):
                cell = ws.cell(row=1, column=col_index)
                cell.value = header
                cell.font = Font(bold=True)
        else:
            all_headers = existing_headers

        # Map header to column index
        header_map = {header: idx + 1 for idx, header in enumerate(all_headers)}

        # Write today's date
        ws.cell(row=delta_days, column=header_map['Date']).value = today_str

        # Write prices to correct columns
        for year in price_columns:
            col_index = header_map.get(year)
            if col_index:
                value = row[year]
                cell = ws.cell(row=delta_days, column=col_index)
                cell.value = value
                cell.number_format = '"$"#,##0.00'

        wb.save(output_file)

    except Exception as e:

        logging.exception(f"🚨 Error during request for {sheet_name}: {e}")

def main():
    # URLs for the different cars, sorted by car make
    urls = {
        'R8': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&sortBy=derivedpriceASC&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&engineCode=10CLDR&makeCode=AUDI&modelCode=R8&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'RS4': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?marketExtension=include&numRecords=24&searchRadius=0&showAccelerateBanner=false&sortBy=derivedpriceASC&zip=32216&allListingType=all-cars&makeCode=AUDI&modelCode=RS4&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'S4': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?marketExtension=include&numRecords=24&searchRadius=500&showAccelerateBanner=false&sortBy=derivedpriceASC&startYear=2013&zip=32216&allListingType=all-cars&makeCode=AUDI&modelCode=S4&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Q5':'https://www.autotrader.com/rest/lsc/crawl/modelyears?mileage=75000&searchRadius=0&sortBy=derivedpriceASC&zip=32216&allListingType=all-cars&makeCode=AUDI&modelCode=Q5&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'RS7':'https://www.autotrader.com/rest/lsc/crawl/modelyears?marketExtension=include&mileage=75000&numRecords=24&searchRadius=0&sortBy=derivedpriceASC&zip=02861&allListingType=all-cars&makeCode=AUDI&modelCode=AUDIRS7&city=Pawtucket&state=RI&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&facetMinCount=3&channel=ATC',

        'M3': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?marketExtension=include&numRecords=24&searchRadius=0&sortBy=derivedpriceASC&transmissionCode=MAN&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&engineCode=8CLDR&makeCode=BMW&modelCode=M3&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'M5': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?endYear=2003&marketExtension=include&numRecords=24&searchRadius=0&sortBy=derivedpriceASC&startYear=2000&zip=32216&allListingType=all-cars&extColorSimple=SILVER&makeCode=BMW&modelCode=M5&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&facetMinCount=3&channel=ATC',

        'Corvette':'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&makeCode=CHEV&modelCode=CORV&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        
        '458 Italia': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&makeCode=FER&modelCode=458ITALIA&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',

        'F250': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=FORD&modelCode=F250&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&facetMinCount=3&channel=ATC',
        'GT':'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=FORD&modelCode=GT&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        
        'S2000': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=HONDA&modelCode=S2000&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&facetMinCount=3&channel=ATC',

        'Wrangler': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=JEEP&modelCode=WRANGLER&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&facetMinCount=3&channel=ATC',

        'Aventador': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=LAM&modelCode=AVENT&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Gallardo': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&sortBy=derivedpriceASC&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&makeCode=LAM&modelCode=GALLARDO&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Huracan': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=LAM&modelCode=LAMHURACAN&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&facetMinCount=3&channel=ATC',
        'Murcielago': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&sortBy=derivedpriceASC&startYear=2007&zip=32216&allListingType=all-cars&makeCode=LAM&modelCode=MURCIELAGO&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Urus': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=LAM&modelCode=LAMURUS&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',

        'AMG':'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&makeCode=MB&modelCode=MBAMGGT&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'C63 AMG': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=MB&seriesCode=C_CLASS&modelCode=C63AMG&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',

        'Carrera S': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?marketExtension=include&mileage=75000&numRecords=24&searchRadius=0&sortBy=derivedpriceASC&startYear=2006&transmissionCode=MAN&zip=32216&allListingType=all-cars&vehicleStyleCode=COUPE&makeCode=POR&modelCode=911&trimCode=911%7CCarrera+S&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'GT3': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?endYear=2024&firstRecord=0&marketExtension=include&numRecords=100&searchRadius=0&sortBy=derivedpriceASC&startYear=2004&zip=32216&makeCode=POR&modelCode=911&trimCode=911%7CGT3&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D&listingType=USED',
        'GT3RS': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=POR&modelCode=911&trimCode=911%7CGT3+RS&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'GT4': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=POR&modelCode=POR718CAY&trimCode=POR718CAY%7CGT4&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'GT4RS': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=POR&modelCode=POR718CAY&trimCode=POR718CAY%7CGT4+RS&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Cayenne': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&makeCode=POR&modelCode=CAYENNE&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Macan': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&zip=32216&allListingType=all-cars&engineCode=6CLDR&makeCode=POR&modelCode=PORMACAN&city=Jacksonville&state=FL&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',

        'Model 3': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&sortBy=derivedpriceASC&zip=02861&allListingType=all-cars&makeCode=TESLA&modelCode=TESMOD3&trimCode=TESMOD3%7CPerformance&city=Pawtucket&state=RI&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D',
        'Model X': 'https://www.autotrader.com/rest/lsc/crawl/modelyears?searchRadius=0&sortBy=derivedpriceASC&zip=02861&allListingType=all-cars&makeCode=TESLA&modelCode=TESMODX&trimCode=TESMODX%7CP100D&city=Pawtucket&state=RI&location=%5Bobject+Object%5D&dma=%5Bobject+Object%5D'
    }

    # Iterate over the URLs and process each one
    for sheet_name, url in urls.items():
        get_avg_price(url, sheet_name)

if __name__ == '__main__':
    main()